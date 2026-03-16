"""
Shared pytest fixtures for excel_table tests.

Reader fixtures are built directly with openpyxl (not via the writer)
so that reader and writer tests remain fully independent.
"""
from __future__ import annotations
from pathlib import Path

import openpyxl
import pytest


# ---------------------------------------------------------------------------
# Common table parameters
# ---------------------------------------------------------------------------

TABLE2D_TITLE        = "Current Map"
TABLE2D_COLUMN_LABEL = "Voltage"
TABLE2D_ROW_LABEL    = "Time"
TABLE2D_COLUMNS      = ["1.0", "2.0", "3.0"]
TABLE2D_ROWS         = ["10.0", "20.0"]
TABLE2D_VALUES       = [[0.1, 0.2, 0.3], [0.4, 0.5, 0.6]]

TABLE1D_TITLE        = "Voltage Points"
TABLE1D_COLUMN_LABEL = "Voltage"
TABLE1D_COLUMNS      = ["1.0", "2.0", "3.0"]
TABLE1D_VALUES       = [[0.1, 0.2, 0.3]]

KV_TITLE   = "Conditions"
KV_COLUMNS = ["temperature", "frequency"]
KV_VALUES  = ["25.0", "50.0"]


# ---------------------------------------------------------------------------
# Table2D fixture
#
# Layout (1-indexed):
#   A1 = title
#   A2:B3 = merged blank (2×2)
#   C2 = column_label
#   C3, D3, E3 = column values
#   A4 = row_label (merged A4:A5)
#   B4, B5 = row values
#   C4:E4, C5:E5 = data values
# ---------------------------------------------------------------------------

@pytest.fixture
def table2d_xlsx(tmp_path: Path) -> Path:
    """Minimal xlsx containing one Table2D written with openpyxl directly."""
    wb = openpyxl.Workbook()
    ws = wb.active

    # title
    ws["A1"] = TABLE2D_TITLE

    # column_label
    ws["C2"] = TABLE2D_COLUMN_LABEL

    # column values
    for ci, val in enumerate(TABLE2D_COLUMNS):
        ws.cell(row=3, column=3 + ci, value=val)

    # row_label
    ws["A4"] = TABLE2D_ROW_LABEL

    # row values and data
    for ri, row_val in enumerate(TABLE2D_ROWS):
        ws.cell(row=4 + ri, column=2, value=row_val)
        for ci, cell_val in enumerate(TABLE2D_VALUES[ri]):
            ws.cell(row=4 + ri, column=3 + ci, value=cell_val)

    path = tmp_path / "table2d.xlsx"
    wb.save(path)
    return path


@pytest.fixture
def table2d_duplicate_xlsx(tmp_path: Path) -> Path:
    """xlsx containing two Table2D blocks with the same title."""
    wb = openpyxl.Workbook()
    ws = wb.active

    def _write(start_row: int):
        ws.cell(row=start_row,     column=1, value=TABLE2D_TITLE)
        ws.cell(row=start_row + 1, column=3, value=TABLE2D_COLUMN_LABEL)
        for ci, val in enumerate(TABLE2D_COLUMNS):
            ws.cell(row=start_row + 2, column=3 + ci, value=val)
        ws.cell(row=start_row + 3, column=1, value=TABLE2D_ROW_LABEL)
        for ri, row_val in enumerate(TABLE2D_ROWS):
            ws.cell(row=start_row + 3 + ri, column=2, value=row_val)
            for ci, cell_val in enumerate(TABLE2D_VALUES[ri]):
                ws.cell(row=start_row + 3 + ri, column=3 + ci, value=cell_val)

    _write(start_row=1)
    _write(start_row=10)

    path = tmp_path / "table2d_duplicate.xlsx"
    wb.save(path)
    return path


# ---------------------------------------------------------------------------
# Table1D fixture
#
# Layout (1-indexed):
#   A1 = title
#   A2 = column_label (merged A2:C2)
#   A3, B3, C3 = column values
#   A4, B4, C4 = data values
# ---------------------------------------------------------------------------

@pytest.fixture
def table1d_xlsx(tmp_path: Path) -> Path:
    """Minimal xlsx containing one Table1D written with openpyxl directly."""
    wb = openpyxl.Workbook()
    ws = wb.active

    ws["A1"] = TABLE1D_TITLE
    ws["A2"] = TABLE1D_COLUMN_LABEL
    for ci, val in enumerate(TABLE1D_COLUMNS):
        ws.cell(row=3, column=1 + ci, value=val)
    for ci, val in enumerate(TABLE1D_VALUES[0]):
        ws.cell(row=4, column=1 + ci, value=val)

    path = tmp_path / "table1d.xlsx"
    wb.save(path)
    return path


# ---------------------------------------------------------------------------
# TableKeyValue fixture
#
# Layout (1-indexed):
#   A1 = title
#   A2, B2 = column (key) values
#   A3, B3 = value strings
# ---------------------------------------------------------------------------

@pytest.fixture
def table_kv_xlsx(tmp_path: Path) -> Path:
    """Minimal xlsx containing one TableKeyValue written with openpyxl directly."""
    wb = openpyxl.Workbook()
    ws = wb.active

    ws["A1"] = KV_TITLE
    for ci, val in enumerate(KV_COLUMNS):
        ws.cell(row=2, column=1 + ci, value=val)
    for ci, val in enumerate(KV_VALUES):
        ws.cell(row=3, column=1 + ci, value=val)

    path = tmp_path / "table_kv.xlsx"
    wb.save(path)
    return path


# ---------------------------------------------------------------------------
# None-value fixture (Table2D with sparse values)
# ---------------------------------------------------------------------------

@pytest.fixture
def table2d_none_xlsx(tmp_path: Path) -> Path:
    """Table2D fixture where some data cells are None (empty)."""
    wb = openpyxl.Workbook()
    ws = wb.active

    ws["A1"] = TABLE2D_TITLE
    ws["C2"] = TABLE2D_COLUMN_LABEL
    for ci, val in enumerate(TABLE2D_COLUMNS):
        ws.cell(row=3, column=3 + ci, value=val)
    ws["A4"] = TABLE2D_ROW_LABEL
    for ri, row_val in enumerate(TABLE2D_ROWS):
        ws.cell(row=4 + ri, column=2, value=row_val)

    # sparse: leave some cells empty
    ws["C4"] = 0.1
    # D4 intentionally empty → None
    ws["E4"] = 0.3
    ws["C5"] = None
    ws["D5"] = 0.5
    ws["E5"] = None

    path = tmp_path / "table2d_none.xlsx"
    wb.save(path)
    return path