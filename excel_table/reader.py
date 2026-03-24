"""
Excel sheet reader.

Locates structured tables within a sheet by scanning for title cells,
then extracts their content into typed Pydantic models.

Usage::

    from excel_table.models import (
        Table2DFloat, Table1DFloat, TableKeyValue,
        FormattedTable2DSchema, FormattedTable1DSchema, TableKeyValueSchema,
    )
    from excel_table.reader import SheetReadSchema, read_sheet

    schema = SheetReadSchema(
        columns=[
            FormattedTable2DSchema(title="Value Map", table_type=Table2DFloat),
            FormattedTable1DSchema(title="x Points", table_type=Table1DFloat,
                                   orientation="vertical"),
            TableKeyValueSchema(title="Conditions"),
        ],
    )
    result = read_sheet("data.xlsx", "Sheet1", schema)
    # result: list[list[...]] — one inner list per detected row of tables
    value_map: Table2DFloat = result[0][0]
"""
from __future__ import annotations

from pathlib import Path
import io

import openpyxl
from pydantic import BaseModel

from .models.table_base import Table1D, Table2D, TableKeyValue
from .models.table_format import (
    FormattedTable1DSchema,
    FormattedTable2DSchema,
    TableKeyValueSchema,
)

ReadSchema = FormattedTable2DSchema | FormattedTable1DSchema | TableKeyValueSchema


class SheetReadSchema(BaseModel):
    """
    Describes the column structure of tables to read from a sheet.

    ``columns`` defines the fixed layout of one logical row of tables:
    the type, table class, and layout metadata for each column position.
    The reader scans the sheet for repeated occurrences of this column
    structure (identified by ``columns[0].title``), reading one row of
    tables at a time until no more are found.

    Attributes:
        columns: List of schema objects defining each column's type and
            layout. ``columns[0].title`` is used as the anchor for detecting
            each row of tables.
        max_scan_rows: Maximum number of rows to scan when searching for
            ``columns[0].title`` (first row) or ``column_label`` in
            ``column_location="bottom"`` tables. Default ``1000``.
    """

    columns: list[ReadSchema]
    max_scan_rows: int = 1000


# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------

def _find_first_row(
    ws,
    title: str,
    max_scan_rows: int,
) -> tuple[int, int] | None:
    """
    Scan the first *max_scan_rows* rows for a cell whose value equals *title*.

    Args:
        ws: openpyxl worksheet object.
        title: Exact string to match.
        max_scan_rows: Maximum number of rows to scan.

    Returns:
        ``(row, col)`` 1-indexed, or ``None`` if not found.
    """
    for row_cells in ws.iter_rows(max_row=max_scan_rows):
        for cell in row_cells:
            if cell.value == title:
                return cell.row, cell.column
    return None


def _find_title_in_row(ws, title: str, row: int) -> int | None:
    """
    Search for *title* along *row*.

    Args:
        ws: openpyxl worksheet object.
        title: Exact string to match.
        row: Row index (1-indexed) to scan.

    Returns:
        Column index (1-indexed) where *title* was found, or ``None``.
    """
    for col_cells in ws.iter_cols(min_row=row, max_row=row):
        cell = col_cells[0]
        if cell.value == title:
            return cell.column
    return None


def _find_title_in_col(ws, title: str, col: int, start_row: int) -> int | None:
    """
    Search for *title* downward in *col* starting from *start_row*.

    Args:
        ws: openpyxl worksheet object.
        title: Exact string to match.
        col: Column index (1-indexed) to scan.
        start_row: Row index (1-indexed) to start scanning from.

    Returns:
        Row index (1-indexed) where *title* was found, or ``None``.
    """
    for row_cells in ws.iter_rows(min_row=start_row, min_col=col, max_col=col):
        cell = row_cells[0]
        if cell.value == title:
            return cell.row
    return None


def _get_merge_span(ws, r: int, c: int) -> tuple[int, int]:
    """
    Return the ``(row_span, col_span)`` of the merged region containing *(r, c)*.

    Returns ``(1, 1)`` if not merged.
    """
    for mr in ws.merged_cells.ranges:
        if mr.min_row <= r <= mr.max_row and mr.min_col <= c <= mr.max_col:
            return mr.max_row - mr.min_row + 1, mr.max_col - mr.min_col + 1
    return 1, 1


def _scan_for_column_label_bottom(
    ws,
    ar: int,
    ac: int,
    row_location: str,
    max_scan_rows: int,
) -> int:
    """
    Scan downward from the row below the title to find ``column_label`` when
    ``column_location="bottom"``.

    ``column_label`` is identified as the first cell whose merged region
    spans >= 2 columns.

    Raises:
        RuntimeError: If not found within *max_scan_rows* rows.
    """
    scan_col = ac + 2 if row_location == "left" else ac
    for offset in range(1, max_scan_rows + 1):
        r = ar + offset
        _, col_span = _get_merge_span(ws, r, scan_col)
        if col_span >= 2:
            return r
    raise RuntimeError(
        f"column_label not found within {max_scan_rows} rows starting at row {ar + 1}. "
        "Check the table format or increase SheetReadSchema.max_scan_rows."
    )


def _read_values_right(ws, r: int, c: int) -> list:
    """Collect contiguous non-``None`` values rightward from *(r, c)*."""
    values = []
    while True:
        val = ws.cell(row=r, column=c).value
        if val is None:
            break
        values.append(val)
        c += 1
    return values


def _read_values_down(ws, r: int, c: int) -> list:
    """Collect contiguous non-``None`` values downward from *(r, c)*."""
    values = []
    while True:
        val = ws.cell(row=r, column=c).value
        if val is None:
            break
        values.append(val)
        r += 1
    return values


# ---------------------------------------------------------------------------
# Table readers
# ---------------------------------------------------------------------------

def _read_table2d(
    ws,
    ar: int,
    ac: int,
    schema: FormattedTable2DSchema,
    max_scan_rows: int,
) -> Table2D:
    col_loc = schema.column_location
    row_loc = schema.row_location

    # column_label
    if col_loc == "top":
        col_label_r = ar + 1
        col_label_c = ac + 2 if row_loc == "left" else ac
    else:
        col_label_r = _scan_for_column_label_bottom(ws, ar, ac, row_loc, max_scan_rows)
        col_label_c = ac + 2 if row_loc == "left" else ac

    _, n_c = _get_merge_span(ws, col_label_r, col_label_c)
    column_label = str(ws.cell(row=col_label_r, column=col_label_c).value)

    # col values
    col_values_r = col_label_r + 1 if col_loc == "top" else col_label_r - 1
    column = [v for v in _read_values_right(ws, col_values_r, col_label_c)]

    # row_label
    if col_loc == "top":
        row_label_r = col_values_r + 1
        row_label_c = ac if row_loc == "left" else ac + n_c + 1
    else:
        row_label_r = ar + 1
        row_label_c = ac if row_loc == "left" else ac + n_c + 1

    n_r, _ = _get_merge_span(ws, row_label_r, row_label_c)
    row_label = str(ws.cell(row=row_label_r, column=row_label_c).value)

    # row values
    row_values_c = row_label_c + 1 if row_loc == "left" else row_label_c - 1
    row = [v for v in _read_values_down(ws, row_label_r, row_values_c)]

    # data matrix
    if col_loc == "top" and row_loc == "left":
        data_r, data_c = row_label_r, ac + 2
    elif col_loc == "top" and row_loc == "right":
        data_r, data_c = row_label_r, ac
    elif col_loc == "bottom" and row_loc == "left":
        data_r, data_c = ar + 1, ac + 2
    else:
        data_r, data_c = ar + 1, ac

    values = [
        [ws.cell(row=data_r + ri, column=data_c + ci).value for ci in range(n_c)]
        for ri in range(n_r)
    ]

    return schema.table_type.model_validate(dict(
        title=ws.cell(row=ar, column=ac).value,
        column_label=column_label,
        row_label=row_label,
        column=column,
        row=row,
        values=values,
    ))


def _read_table1d(
    ws,
    ar: int,
    ac: int,
    schema: FormattedTable1DSchema,
) -> Table1D:
    column_label = str(ws.cell(row=ar + 1, column=ac).value)

    if schema.orientation == "horizontal":
        column = [v for v in _read_values_right(ws, ar + 2, ac)]
        values = [[ws.cell(row=ar + 3, column=ac + ci).value for ci in range(len(column))]]
    else:
        n_c, _ = _get_merge_span(ws, ar + 1, ac)
        column = [str(ws.cell(row=ar + 1 + ci, column=ac + 1).value) for ci in range(n_c)]
        values = [[ws.cell(row=ar + 1 + ci, column=ac + 2).value] for ci in range(n_c)]

    return schema.table_type.model_validate(dict(
        title=ws.cell(row=ar, column=ac).value,
        column_label=column_label,
        column=column,
        values=values,
    ))


def _read_table_key_value(
    ws,
    ar: int,
    ac: int,
    schema: TableKeyValueSchema,
) -> TableKeyValue:
    column = [v for v in _read_values_right(ws, ar + 1, ac)]
    value = [ws.cell(row=ar + 2, column=ac + ci).value for ci in range(len(column))]

    return schema.table_type.model_validate(dict(
        title=ws.cell(row=ar, column=ac).value,
        column=column,
        value=value,
    ))


def _read_one(
    ws,
    ar: int,
    ac: int,
    item: ReadSchema,
    max_scan_rows: int,
) -> Table1D | Table2D | TableKeyValue:
    if isinstance(item, TableKeyValueSchema):
        return _read_table_key_value(ws, ar, ac, item)
    elif isinstance(item, FormattedTable2DSchema):
        return _read_table2d(ws, ar, ac, item, max_scan_rows)
    elif isinstance(item, FormattedTable1DSchema):
        return _read_table1d(ws, ar, ac, item)
    else:
        raise ValueError(f"Unknown schema type: {type(item)}")


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def _read_from_worksheet(ws, sheet_name: str, schema: SheetReadSchema) -> list[list[Table1D | Table2D | TableKeyValue]]:
    """
    Extract structured tables from an already-opened worksheet.

    Extracted from :func:`read_sheet` so that file-path and bytes-based
    entry points can share the same parsing logic.

    Args:
        ws: openpyxl worksheet object.
        sheet_name: Worksheet name, used only in error messages.
        schema: Column structure definition.

    Returns:
        List of rows, each row being a list of validated table instances.
    """
    anchor_title = schema.columns[0].title
    result: list[list] = []
    pos = _find_first_row(ws, anchor_title, schema.max_scan_rows)
    if pos is None:
        return result
    anchor_row, anchor_col = pos
    while True:
        col_positions: list[int] = [anchor_col]
        for item in schema.columns[1:]:
            c = _find_title_in_row(ws, item.title, anchor_row)
            if c is None:
                raise ValueError(
                    f"Column title '{item.title}' not found in row {anchor_row} "
                    f"of sheet '{sheet_name}'"
                )
            col_positions.append(c)
        row_result = [
            _read_one(ws, anchor_row, col, item, schema.max_scan_rows)
            for item, col in zip(schema.columns, col_positions)
        ]
        result.append(row_result)
        next_row = _find_title_in_col(ws, anchor_title, anchor_col, anchor_row + 1)
        if next_row is None:
            break
        anchor_row = next_row
    return result


def read_sheet(
    path: str | Path,
    sheet_name: str,
    schema: SheetReadSchema,
) -> list[list[Table1D | Table2D | TableKeyValue]]:
    """
    Read structured tables from an Excel sheet.

    The reader detects repeated rows of tables by scanning for
    ``schema.columns[0].title``:

    1. The first occurrence is found by scanning up to ``schema.max_scan_rows``
       rows with ``iter_rows``.
    2. From that row, each remaining column title is located in the same row.
    3. For subsequent rows, ``columns[0].title`` is searched downward in the
       same column. Scanning stops when no further occurrence is found.

    Args:
        path: Path to the ``.xlsx`` file.
        sheet_name: Name of the worksheet to read from.
        schema: Column structure definition.

    Returns:
        List of rows, each row being a list of validated table instances
        corresponding to ``schema.columns``.

    Raises:
        KeyError: If *sheet_name* does not exist in the workbook.
        ValueError: If any column title is not found in the expected row,
            or if extracted data fails model validation.
        RuntimeError: If ``column_label`` cannot be located for a
            ``column_location="bottom"`` table.
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet_name]
    return _read_from_worksheet(ws, sheet_name, schema)


def read_sheet_bytes(
    data: bytes,
    sheet_name: str,
    schema: SheetReadSchema,
) -> list[list[Table1D | Table2D | TableKeyValue]]:
    """
    Read structured tables from an Excel file supplied as raw bytes.

    Equivalent to :func:`read_sheet` but accepts ``bytes`` instead of a
    file path. Useful when the workbook is held in memory (e.g. returned
    by :func:`write_sheet_bytes`) without being written to disk.

    Args:
        data: Raw ``.xlsx`` file contents.
        sheet_name: Name of the worksheet to read from.
        schema: Column structure definition.

    Returns:
        List of rows, each row being a list of validated table instances
        corresponding to ``schema.columns``.

    Raises:
        KeyError: If *sheet_name* does not exist in the workbook.
        ValueError: If any column title is not found in the expected row,
            or if extracted data fails model validation.
        RuntimeError: If ``column_label`` cannot be located for a
            ``column_location="bottom"`` table.
    """
    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    ws = wb[sheet_name]
    return _read_from_worksheet(ws, sheet_name, schema)