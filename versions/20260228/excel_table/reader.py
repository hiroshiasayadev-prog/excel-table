"""
Excel sheet reader.

Locates structured tables within a sheet by scanning for title cells,
then extracts their content into typed Pydantic models.

Usage::

    schema = SheetReadSchema(rows=[
        [("Value Map", Table2DFloat)],
        [("Operating Conditions", TableKeyValue)],
    ])
    result = read_sheet("data.xlsx", "Sheet1", schema)
    value_map: Table2DFloat = result[0][0]
"""
from __future__ import annotations
from pathlib import Path
from typing import Type

import openpyxl
from pydantic import BaseModel

from .models.table import Table1D, Table2D, TableKeyValue

TableType = Type[Table1D | Table2D | TableKeyValue]


class SheetReadSchema(BaseModel):
    """
    Describes which tables to read from a sheet and their expected types.

    ``rows`` mirrors the logical layout of tables in the sheet: each inner list
    is one horizontal group of tables. The order within each group determines
    the order of results returned by :func:`read_sheet`.

    Attributes:
        rows: Nested list of ``(title, TableType)`` pairs. ``TableType`` may be
            any of the typed subclasses (e.g. ``Table2DFloat``) or a domain
            subclass, as long as it inherits from :class:`Table1D`, :class:`Table2D`,
            or :class:`TableKeyValue`. Pydantic ``model_validate`` is called with
            the provided type, so coercion rules of that type apply.
    """

    rows: list[list[tuple[str, TableType]]]


def _find_cell(ws, title: str) -> tuple[int, int] | None:
    """
    Scan the worksheet for a cell whose value equals *title*.

    Merged cells: openpyxl returns ``None`` for all cells in a merged range
    except the top-left, so this scan correctly identifies the anchor cell.

    Args:
        ws: openpyxl worksheet object.
        title: Exact string to match against cell values.

    Returns:
        ``(row, col)`` 1-indexed, or ``None`` if not found.
    """
    for row in ws.iter_rows():
        for cell in row:
            if cell.value == title:
                return cell.row, cell.column
    return None


def _read_until_none(ws, start_row: int, start_col: int, direction: str) -> list:
    """
    Collect contiguous cell values in one direction, stopping at the first ``None``.

    Args:
        ws: openpyxl worksheet object.
        start_row: Starting row (1-indexed).
        start_col: Starting column (1-indexed).
        direction: ``"right"`` to scan columns, ``"down"`` to scan rows.

    Returns:
        List of non-``None`` cell values in order.
    """
    values = []
    r, c = start_row, start_col
    while True:
        val = ws.cell(row=r, column=c).value
        if val is None:
            break
        values.append(val)
        if direction == "right":
            c += 1
        else:
            r += 1
    return values


def _read_table2d(ws, r: int, c: int, table_type: type) -> Table2D:
    """
    Read a :class:`Table2D` from worksheet *ws* anchored at *(r, c)*.

    Expected layout (1-indexed offsets from anchor)::

        (r+0, c+0) title [single cell]
        (r+1, c+0) [2×2 merged blank]     (r+1, c+2) column_label [merged]
        (r+2, c+0)                         (r+2, c+2) col1  col2 ...
        (r+3, c+0) row_label [merged down] (r+3, c+1) row1  val  ...
        (r+4, c+0)                         (r+4, c+1) row2  val  ...

    Args:
        ws: openpyxl worksheet object.
        r: Anchor row (1-indexed), where the title cell is located.
        c: Anchor column (1-indexed).
        table_type: Concrete subclass to instantiate via ``model_validate``.

    Returns:
        Validated instance of *table_type*.
    """
    column_label = str(ws.cell(row=r + 1, column=c + 2).value)
    column = [str(v) for v in _read_until_none(ws, r + 2, c + 2, "right")]
    row_label = str(ws.cell(row=r + 3, column=c).value)
    row = [str(v) for v in _read_until_none(ws, r + 3, c + 1, "down")]

    values = []
    for ri in range(len(row)):
        row_vals = []
        for ci in range(len(column)):
            row_vals.append(ws.cell(row=r + 3 + ri, column=c + 2 + ci).value)
        values.append(row_vals)

    return table_type.model_validate(
        dict(
            title=ws.cell(row=r, column=c).value,
            column_label=column_label,
            row_label=row_label,
            column=column,
            row=row,
            values=values,
        )
    )


def _read_table1d(ws, r: int, c: int, table_type: type) -> Table1D:
    """
    Read a :class:`Table1D` from worksheet *ws* anchored at *(r, c)*.

    Args:
        ws: openpyxl worksheet object.
        r: Anchor row (1-indexed).
        c: Anchor column (1-indexed).
        table_type: Concrete subclass to instantiate via ``model_validate``.

    Returns:
        Validated instance of *table_type*.
    """
    column_label = str(ws.cell(row=r + 1, column=c).value)
    column = [str(v) for v in _read_until_none(ws, r + 2, c, "right")]
    row_vals = [ws.cell(row=r + 3, column=c + ci).value for ci in range(len(column))]

    return table_type.model_validate(
        dict(
            title=ws.cell(row=r, column=c).value,
            column_label=column_label,
            column=column,
            values=[row_vals],
        )
    )


def _read_table_key_value(ws, r: int, c: int, table_type: type) -> TableKeyValue:
    """
    Read a :class:`TableKeyValue` from worksheet *ws* anchored at *(r, c)*.

    Args:
        ws: openpyxl worksheet object.
        r: Anchor row (1-indexed).
        c: Anchor column (1-indexed).
        table_type: Concrete subclass to instantiate via ``model_validate``.

    Returns:
        Validated instance of *table_type*.
    """
    column = [str(v) for v in _read_until_none(ws, r + 1, c, "right")]
    value = [
        str(ws.cell(row=r + 2, column=c + ci).value) for ci in range(len(column))
    ]
    return table_type.model_validate(
        dict(
            title=ws.cell(row=r, column=c).value,
            column=column,
            value=value,
        )
    )


def read_sheet(
    path: str | Path,
    sheet_name: str,
    schema: SheetReadSchema,
) -> list[list[Table1D | Table2D | TableKeyValue]]:
    """
    Read structured tables from an Excel sheet.

    Tables are located by scanning the sheet for cells matching each title in
    *schema*. The sheet is opened in ``data_only=True`` mode so formula cells
    return their cached values.

    Args:
        path: Path to the ``.xlsx`` file.
        sheet_name: Name of the worksheet to read from.
        schema: Describes which tables to read and their expected types.

    Returns:
        Nested list mirroring ``schema.rows``, with each ``(title, type)`` pair
        replaced by the corresponding validated table instance.

    Raises:
        KeyError: If *sheet_name* does not exist in the workbook.
        ValueError: If a title from the schema is not found in the sheet,
            or if the extracted data fails model validation.
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet_name]

    result: list[list] = []
    for schema_row in schema.rows:
        row_result = []
        for title, table_type in schema_row:
            pos = _find_cell(ws, title)
            if pos is None:
                raise ValueError(
                    f"Table with title '{title}' not found in sheet '{sheet_name}'"
                )
            r, c = pos

            if issubclass(table_type, TableKeyValue):
                table = _read_table_key_value(ws, r, c, table_type)
            elif issubclass(table_type, Table2D):
                table = _read_table2d(ws, r, c, table_type)
            elif issubclass(table_type, Table1D):
                table = _read_table1d(ws, r, c, table_type)
            else:
                raise ValueError(f"Unknown table type: {table_type}")

            row_result.append(table)
        result.append(row_result)

    return result